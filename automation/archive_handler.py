import credentials
from database import connection as cnn
from tkinter import filedialog
import PySimpleGUI as sg
import openpyxl
import os
import adapter

def abrir_arquivo(initialdir = 'C:/Users/Licitacao/Desktop',filetypes=(('Arquivo','*.*'),)):
    return filedialog.askopenfilename(initialdir=initialdir,filetypes=filetypes)

def abrir_pasta(initialdir='C:/Users/Licitacao/Desktop'):
    return filedialog.askdirectory(initialdir=initialdir)

class Planilha:
    
    def __init__(self, caminho):
        self.planilha = openpyxl.load_workbook(caminho,data_only=True)

    def obter_pregao(self):
        wb = self.planilha['Controle']
        pregao = {
            'numero' : adapter.padronizar_pregao(str(wb.cell(2,1).value)),
            'uasg' : adapter.padronizar_uasg(str(wb.cell(2,2).value)),
            'data' : adapter.padronizar_data_hora(str(wb.cell(2,3).value)[:10],str(wb.cell(2,4).value)[:5]),
            'orgao' : str(wb.cell(2,5).value).upper()
        }
        return pregao
    
    def obter_itens_cotados(self):
        wb = self.planilha['Planilha1']
        colunas_registro = [1,3,4,5,7,9,11,12,13]
        item = []
        for linha in range(2,wb.max_row):
            item.append([str(wb.cell(linha,coluna).value) for coluna in colunas_registro])
        return item

def renomear_arquivo(pasta:str, arquivo:str, nomenclatura:str):
    """Renomea arquivo de acordo com a nomenclatura inserida."""
    while True:
        try:
            os.rename(pasta +'/'+ arquivo, pasta +'/'+ nomenclatura +'_'+ arquivo)
            break
        except:
            sg.popup('Favor fechar os arquivos do processo para renomea-los.')

def mover_e_renomear_pasta(pasta:str,nomenclatura:str):
    local = '/'.join(pasta.split('/')[:-1])+'/'
    while True:
        try:
            os.rename(pasta,local+nomenclatura)
            break
        except:
            sg.popup('Favor fechar qualquer documento que encontre-se dentro da pasta.')

def cadastrar_planilha(renomear:bool):
    """Realiza a leitura da planilha de cotação, insere os dados em banco de dados e move a pasta para a pasta de pregões"""
    try:
        pasta = abrir_pasta()
        pasta_proposta = pasta+'\proposta'
        for arquivo in os.listdir(pasta_proposta):
            if arquivo.endswith('.xlsx'):
                arquivo_planilha = arquivo
            if arquivo.endswith('.docx'):
                arquivo_word = arquivo
    except:
        sg.popup('Não foi possível abrir o arquivo.')
        return
    try:
        planilha = Planilha(pasta_proposta+'/'+arquivo_planilha)
    except:
        sg.popup('Confira se a planilha está aberta, caso esteja favor fechar.')
        return
    pregao = planilha.obter_pregao()
    itens = planilha.obter_itens_cotados()
    if(cnn.consultar_id_orgao(pregao['uasg'])!='-1'):
        if not cnn.inserir_pregao(pregao['uasg'],pregao['numero'],pregao['data'],"Proposta"):
            sg.popup('O pregão já foi inserido anteriormente.')
            return
    else:
        cnn.inserir_orgao(pregao['uasg'],pregao['orgao'])
        cnn.inserir_pregao(pregao['uasg'],pregao['numero'],pregao['data'],"Proposta")

    aux=[]
    for item in itens:
        aux.append(dict(
            item=item[0],
            modelo=item[7],
            quantidade=item[3],
            valor_ofertado=item[2],
            preco_custo=item[4],
            frete=item[5],
            fornecedor=item[8],
            marca=item[1],
            categoria=item[6]))
    if cnn.inserir_itens_planilha(pregao['uasg'],pregao['numero'],aux):
        if (renomear):
            nomenclatura = adapter.padronizar_nome_pasta(pregao['data'],pregao['numero'],pregao['uasg'])
            renomear_arquivo(pasta_proposta, arquivo_planilha, nomenclatura)
            renomear_arquivo(pasta_proposta, arquivo_word, nomenclatura)
            mover_e_renomear_pasta(pasta,nomenclatura)
        sg.popup('Foram inseridos '+str(len(itens))+' itens no pregão de número '+pregao['numero'])
    else:
        sg.popup('Ocorreu um erro ao tentar inserir um dos itens do pregão.')